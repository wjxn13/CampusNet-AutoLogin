import pyautogui
import time
import os
from openpyxl import load_workbook
from pyautogui import ImageNotFoundException

print("当前工作目录：", os.getcwd())

pyautogui.FAILSAFE = True
IMAGE_DIR = "images"

# 分步骤使用的置信度
NORMAL_CONFIDENCE = 0.75      # 其他步骤（匹配精准）
WLAN_ICON_CONFIDENCE = 0.6    # 网络图标（匹配度较低）


def load_tasks(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    tasks = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        step, image, action, value, required, *_ = row
        tasks.append({
            "image": image,
            "action": action,
            "value": value,
            "required": str(required).upper() == "是"
        })
    return tasks


def wait_and_click(image_name, confidence, timeout=5, click_offset=None):
    """
    在 timeout 秒内反复查找图片，找到后立即点击，返回 True/False
    click_offset: None（点击中心）、("right", 比例) 或 ("center",)
    """
    path = os.path.join(IMAGE_DIR, image_name)
    if not os.path.exists(path):
        print(f"❌ 图片不存在: {path}")
        return False

    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            region = pyautogui.locateOnScreen(path, confidence=confidence)
        except ImageNotFoundException:
            region = None

        if region:
            left, top, width, height = region
            if click_offset and click_offset[0] == "right":
                click_x = left + width * click_offset[1]
                click_y = top + height * 0.5
            else:
                click_x, click_y = pyautogui.center(region)

            pyautogui.click(click_x, click_y)
            print(f"   🖱️ 点击 ({click_x:.0f}, {click_y:.0f})")
            return True

        time.sleep(0.3)   # 轮询间隔，既快又不给 CPU 太大压力

    return False


def do_step(task):
    # ---------- 特殊处理：网络图标（两种状态，低置信度） ----------
    if task["image"] == "wlan_icon.png":
        icon_candidates = ["wlan_icon.png", "wlan_disconnected.png"]

        for icon_name in icon_candidates:
            print(f"🔍 尝试查找网络图标：{icon_name}")
            if wait_and_click(icon_name, confidence=WLAN_ICON_CONFIDENCE, timeout=3):
                print(f"   ✅ 点击网络图标：{icon_name}")
                time.sleep(0.3)   # 微小延迟，确保点击生效
                return True

        print("❌ 未找到任何网络图标")
        return False

    # ---------- 其他步骤统一处理 ----------
    print(f"🔍 正在查找图片：{task['image']}")

    # 根据图片名配置不同的超时和点击方式
    if task["image"] == "wlan_arrow.png":
        success = wait_and_click("wlan_arrow.png", confidence=NORMAL_CONFIDENCE,
                                 timeout=5, click_offset=("right", 0.8))
    elif task["image"] == "campus_wifi.png":
        success = wait_and_click("campus_wifi.png", confidence=NORMAL_CONFIDENCE, timeout=5)
    elif task["image"] == "connect_btn.png":
        if wait_and_click("connect_btn.png", confidence=NORMAL_CONFIDENCE, timeout=3):
            print("   ⏳ 已点击连接，等待网络响应...")
            time.sleep(25)          # 网络连接耗时不可缩短
            return True
        else:
            success = False
    elif task["image"] == "login_page.png":
        success = wait_and_click("login_page.png", confidence=NORMAL_CONFIDENCE, timeout=5)
    elif task["image"] == "login_btn.png":
        success = wait_and_click("login_btn.png", confidence=NORMAL_CONFIDENCE, timeout=3)
    elif task["image"] in ["username_input.png", "password_input.png"]:
        # 输入框处理（目前只做了点击，如需输入可扩展）
        success = wait_and_click(task["image"], confidence=NORMAL_CONFIDENCE, timeout=2)
    else:
        # 其他未知图片（如 browser_hint.png）
        success = wait_and_click(task["image"], confidence=NORMAL_CONFIDENCE, timeout=3)

    if success:
        time.sleep(0.2)   # 微小保护延迟
        return True
    else:
        print(f"❌ 未找到图片：{task['image']}")
        return False


def main():
    print("5 秒后开始，请不要操作鼠标")
    time.sleep(5)

    tasks = load_tasks("config.xlsx")
    valid_tasks = [t for t in tasks if t["image"]]
    print(f"共加载 {len(valid_tasks)} 个有效步骤")

    skip_input_steps = False

    for i, task in enumerate(valid_tasks, 1):
        print(f"\n{'=' * 50}")
        print(f"步骤 {i}/{len(valid_tasks)}：{task['image']} ({'必选' if task['required'] else '可选'})")
        print(f"{'=' * 50}")

        # 登录页面出现时，判断是否已记住密码
        if task["image"] == "login_page.png":
            print("🔍 检测是否已记住密码...")
            username_path = os.path.join(IMAGE_DIR, "username_input.png")
            try:
                region = pyautogui.locateOnScreen(username_path, confidence=0.75)
                if region:
                    print("   📝 未记住密码，需要输入账号")
                    skip_input_steps = False
                else:
                    print("   ✅ 已记住密码，直接登录")
                    skip_input_steps = True
            except ImageNotFoundException:
                print("   ✅ 已记住密码，直接登录")
                skip_input_steps = True

        # 跳过输入步骤
        if skip_input_steps and task["image"] in ["username_input.png", "password_input.png"]:
            print(f"⏭️ 跳过输入步骤：{task['image']}")
            continue

        ok = do_step(task)
        if not ok:
            if task["required"]:
                print(f"\n❌ 必选步骤失败，脚本终止！")
                raise RuntimeError(f"未找到必选图片: {task['image']}")
            else:
                print(f"⚠️ 跳过可选步骤: {task['image']}")

        time.sleep(0.2)   # 步骤间极短间隔

    print("\n🎉 校园网登录完成！")


if __name__ == "__main__":
    main()