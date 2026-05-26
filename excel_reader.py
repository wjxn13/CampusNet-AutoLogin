from openpyxl import load_workbook

def load_tasks(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    tasks = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        image, action, value, _ = row
        tasks.append((image, action, value))
    return tasks